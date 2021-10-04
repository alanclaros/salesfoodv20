from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
# settings de la app
from django.conf import settings
# reverse url
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.apps import apps

# propios
from configuraciones.models import Cajas, Puntos, Sucursales, Ciudades, Almacenes, Lineas
from permisos.models import UsersPerfiles
from cajas.models import CajasIngresos, CajasEgresos
from status.models import Status

# para los usuarios
from utils.permissions import get_user_permission_operation, get_permissions_user, get_html_column
from utils.dates_functions import get_date_system, get_date_show, get_date_to_db

# clases por modulo
from controllers.reportes.ReportesController import ReportesController

import os
# xls
import openpyxl

# reportes
import io
from django.http import FileResponse, HttpResponse
from reportlab.pdfgen import canvas

from reportes.rptArqueoCaja import rptArqueoCaja
from reportes.rptIngresosCaja import rptIngresosCaja
from reportes.rptEgresosCaja import rptEgresosCaja
from reportes.rptMovimientosCaja import rptMovimientosCaja
from reportes.rptIngresoProductos import rptIngresoProductos
from reportes.rptSalidaProductos import rptSalidaProductos
from reportes.rptMovimientoProductos import rptMovimientoProductos
from reportes.rptPedidosProductos import rptPedidosProductos
from reportes.rptPreventas import rptPreventas
from reportes.rptVentas import rptVentas
from reportes.rptPlanPagos import rptPlanPagos
from reportes.rptStockProductos import rptStockProductos

# reportes index


@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def reportes_index(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = reportes_controller.status_activo

    # operaciones
    if 'operation_x' in request.POST.keys():
        # verificamos operacion valida
        operation = request.POST['operation_x']
        if not operation in ['', 'arqueo_caja', 'ingresos_caja', 'egresos_caja', 'movimientos_caja', 'buscar_sucursal', 'buscar_solo_sucursal', 'buscar_sucursal_punto', 'buscar_caja', 'buscar_punto',
                             'ingreso_productos', 'buscar_sucursal_almacen', 'buscar_almacen', 'salida_productos', 'movimiento_productos', 'pedido_productos',
                             'preventas', 'ventas', 'plan_pagos', 'stock_productos']:

            return render(request, 'pages/without_permission.html')

        # stock productos
        if operation == 'stock_productos':
            respuesta = stock_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        # plan pagos
        if operation == 'plan_pagos':
            respuesta = plan_pagos(request)
            if not type(respuesta) == bool:
                return respuesta

        # ventas
        if operation == 'ventas':
            respuesta = ventas(request)
            if not type(respuesta) == bool:
                return respuesta

        # preventas
        # if operation == 'preventas':
        #     url = reverse('rpt_preventas')
        #     return HttpResponseRedirect(url)

        # inventarios
        if operation == 'ingreso_productos':
            respuesta = ingreso_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'salida_productos':
            respuesta = salida_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'movimiento_productos':
            respuesta = movimiento_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'pedido_productos':
            respuesta = pedido_productos(request)
            if not type(respuesta) == bool:
                return respuesta

        # cajas
        if operation == 'movimientos_caja':
            respuesta = movimientos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'egresos_caja':
            respuesta = egresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'ingresos_caja':
            respuesta = ingresos_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'arqueo_caja':
            respuesta = arqueo_caja(request)
            if not type(respuesta) == bool:
                return respuesta

        if operation == 'buscar_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal.html', context_s)

        if operation == 'buscar_solo_sucursal':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_solo_sucursal.html', context_s)

        if operation == 'buscar_sucursal_punto':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal_punto.html', context_s)

        if operation == 'buscar_sucursal_almacen':
            ciudad_id = request.POST['ciudad'].strip()
            ciudad = Ciudades.objects.get(pk=ciudad_id)
            lista_sucursales = Sucursales.objects.filter(status_id=status_activo, ciudad_id=ciudad).order_by('sucursal')

            context_s = {
                'lista_sucursales': lista_sucursales,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_sucursal_almacen.html', context_s)

        if operation == 'buscar_caja':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['punto_id__sucursal_id'] = sucursal

            lista_cajas = Cajas.objects.select_related('punto_id').filter(**filtro).order_by('punto_id__punto', 'caja')

            context_s = {
                'lista_cajas': lista_cajas,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_caja.html', context_s)

        if operation == 'buscar_punto':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['sucursal_id'] = sucursal

            lista_puntos = Puntos.objects.select_related('sucursal_id').filter(**filtro).order_by('punto')

            context_s = {
                'lista_puntos': lista_puntos,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_punto.html', context_s)

        if operation == 'buscar_almacen':
            sucursal_id = request.POST['sucursal'].strip()
            sucursal = Sucursales.objects.get(pk=sucursal_id)
            filtro = {}
            filtro['status_id'] = status_activo
            filtro['sucursal_id'] = sucursal

            lista_almacenes = Almacenes.objects.select_related('sucursal_id').filter(**filtro).order_by('almacen')

            context_s = {
                'lista_almacenes': lista_almacenes,
                'autenticado': 'si',
            }
            return render(request, 'reportes/busqueda_almacen.html', context_s)

    # verificamos mensajes
    if 'nuevo_mensaje' in request.session.keys():
        messages.add_message(request, messages.SUCCESS, request.session['nuevo_mensaje'])
        del request.session['nuevo_mensaje']
        request.session.modified = True

    # datos por defecto
    context = {
        'permisos': permisos,
        'url_main': '',
        # 'js_file': reportes_controller.modulo_session,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': '',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/reportes.html', context)


# arqueo de caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def arqueo_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja
        fecha = get_date_to_db(fecha=request.POST['fecha'], formato_ori='dd-MMM-yyyy', formato='yyyy-mm-dd')

        if permisos.imprimir and reportes_controller.verificar_permiso(request.user, caja_id=int(request.POST['id'])):
            buffer = io.BytesIO()
            rptArqueoCaja(buffer, int(request.POST['id']), fecha)

            buffer.seek(0)
            return FileResponse(buffer, filename='arqueo_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'arqueo_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/arqueo_caja.html', context)


# ingresos a caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ingresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptIngresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            return FileResponse(buffer, filename='ingresos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_ingreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ingresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas de la sucursal
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'fecha_actual': fecha_actual,
        'lista_ciudades': lista_ciudades,
        'permisos': permisos,
        'url_actual': '',

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ingresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ingresos_caja.html', context)


# egresos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def egresos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptEgresosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='egresos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Caja', 'Operacion', 'Fecha', 'concepto', 'monto', 'codigo', 'estado'))
            datos_reporte = reportes_controller.datos_egreso_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['caja'], dato['operacion'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=egresos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista ciudades
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    # lista de cajas
    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'egresos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/egresos_caja.html', context)


# movimientos caja
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def movimientos_caja(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptMovimientosCaja(buffer, request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimientos_caja.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        caja = int(request.POST['caja'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Valores"
            hoja.append(('Ciudad', 'Sucursal', 'Fecha', 'Concepto', 'Monto', 'Codigo', 'Caja1', 'Punto1', 'Caja2', 'Punto2', 'estado'))
            datos_reporte = reportes_controller.datos_movimientos_caja(request.user, ciudad, sucursal, caja, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['fecha'], dato['concepto'], dato['monto'], dato['tipo_moneda'], dato['caja1'], dato['punto1'], dato['caja2'], dato['punto2'], dato['estado_txt']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimientos_caja.xlsx"
            wb.save(response)
            return response
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_cajas = reportes_controller.lista_cajas_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_cajas': lista_cajas,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'movimientos_caja',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/movimientos_caja.html', context)


# ingreso de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ingreso_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptIngresoProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='ingreso_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "IngresoProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Fecha Elaboracion', 'Fecha Vencimiento', 'Lote', 'Cantidad'))
            datos_reporte = reportes_controller.datos_ingreso_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['fecha_elaboracion'], dato['fecha_vencimiento'], dato['lote'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ingreso_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ingreso_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ingreso_productos.html', context)


# salida de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def salida_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptSalidaProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='salida_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "SalidaProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Fecha Elaboracion', 'Fecha Vencimiento', 'Lote', 'Cantidad'))
            datos_reporte = reportes_controller.datos_salida_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['fecha_elaboracion'], dato['fecha_vencimiento'], dato['lote'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=salida_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'salida_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/salida_productos.html', context)


# movimiento de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def movimiento_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptMovimientoProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimiento_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "MovimientoProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Fecha Elaboracion', 'Fecha Vencimiento', 'Lote', 'Cantidad'))
            datos_reporte = reportes_controller.datos_movimiento_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['fecha_elaboracion'], dato['fecha_vencimiento'], dato['lote'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimiento_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'movimiento_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/movimiento_productos.html', context)


# pedido de productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def pedido_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptPedidosProductos(buffer, request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='movimiento_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        almacen = int(request.POST['almacen'].strip())
        anulados = request.POST['anulados'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "PedidoProductos"
            hoja.append(('Ciudad', 'Sucursal', 'Almacen', 'Linea', 'Producto', 'Fecha Elaboracion', 'Fecha Vencimiento', 'Lote', 'Cantidad'))
            datos_reporte = reportes_controller.datos_pedido_productos(request.user, ciudad, sucursal, almacen, fecha_ini, fecha_fin, anulados)
            for dato in datos_reporte:
                hoja.append((dato['ciudad'], dato['sucursal'], dato['almacen'], dato['linea'], dato['producto'], dato['fecha_elaboracion'], dato['fecha_vencimiento'], dato['lote'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=movimiento_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_almacenes': lista_almacenes,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'pedido_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/pedido_productos.html', context)


# # preventas
# @user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
# def preventas(request):
#     # url main
#     url_main = reverse('reportes')
#     permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

#     # controler
#     reportes_controller = ReportesController()
#     status_activo = Status.objects.get(pk=reportes_controller.activo)

#     if 'operation_x' in request.POST.keys() and request.POST['operation_x'] == 'print':
#         fecha_ini = request.POST['fecha_ini'].strip()
#         fecha_fin = request.POST['fecha_fin'].strip()
#         ciudad = int(request.POST['ciudad'].strip())
#         sucursal = int(request.POST['sucursal'].strip())
#         punto = int(request.POST['punto'].strip())
#         anulados = request.POST['anulados'].strip()
#         inventario = request.POST['inventario'].strip()
#         venta = request.POST['venta'].strip()

#         if permisos.imprimir:
#             buffer = io.BytesIO()
#             rptPreventas(buffer, request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, inventario=inventario, venta=venta)

#             buffer.seek(0)
#             # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
#             return FileResponse(buffer, filename='preventas.pdf')
#         else:
#             url = reverse('without_permission')
#             return HttpResponseRedirect(url)

#     if 'operation_x' in request.POST.keys() and request.POST['operation_x'] == 'print_excel':
#         # reporte de arqueo de caja

#         fecha_ini = request.POST['fecha_ini'].strip()
#         fecha_fin = request.POST['fecha_fin'].strip()
#         ciudad = int(request.POST['ciudad'].strip())
#         sucursal = int(request.POST['sucursal'].strip())
#         punto = int(request.POST['punto'].strip())
#         anulados = request.POST['anulados'].strip()
#         inventario = request.POST['inventario'].strip()
#         venta = request.POST['venta'].strip()

#         if permisos.imprimir:
#             buffer = io.BytesIO()
#             buffer.seek(0)

#             wb = openpyxl.Workbook()
#             hoja = wb.active
#             hoja.title = "Preventas"
#             hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Cliente', 'Tipo', 'Subtotal', 'Descuento', 'Total', 'Numero', 'Estado', 'Fecha'))
#             datos_reporte = reportes_controller.datos_preventas(request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, inventario=inventario, venta=venta)
#             for dato in datos_reporte:
#                 estado = 'preventa'
#                 if dato['status_id'] == reportes_controller.preventa_venta:
#                     estado = 'venta'
#                 if dato['status_id'] == reportes_controller.anulado:
#                     estado = 'anulado'

#                 hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['cliente'], dato['tipo'], dato['subtotal'], dato['descuento'], dato['total'], dato['numero'], estado, dato['fecha']))

#             response = HttpResponse(content_type="application/msexcel")
#             response["Content-Disposition"] = "attachment; filename=preventas.xlsx"
#             wb.save(response)
#             return response

#     # lista de cajas de la sucursal
#     lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

#     lista_puntos = reportes_controller.lista_puntos_sucursal(request.user)
#     fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

#     # datos por defecto
#     context = {
#         'url_main': url_main,
#         'js_file': reportes_controller.modulo_session,
#         'lista_puntos': lista_puntos,
#         'lista_ciudades': lista_ciudades,
#         'fecha_actual': fecha_actual,
#         'permisos': permisos,

#         'autenticado': 'si',
#     }
#     return render(request, 'reportes/preventas.html', context)


# ventas
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def ventas(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        tipo = request.POST['tipo'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptVentas(buffer, request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, tipo=tipo)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='ventas.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        tipo = request.POST['tipo'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Ventas"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Cliente', 'Tipo', 'Subtotal', 'Descuento', 'Total', 'Numero', 'Estado', 'Fecha'))
            datos_reporte = reportes_controller.datos_ventas(request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, tipo=tipo)
            for dato in datos_reporte:
                estado = 'contado'
                if dato['status_id'] == reportes_controller.contado:
                    estado = 'contado'
                if dato['status_id'] == reportes_controller.factura:
                    estado = 'factura'
                if dato['status_id'] == reportes_controller.consignacion:
                    estado = 'consignacion'
                if dato['status_id'] == reportes_controller.plan_pago:
                    estado = 'planpago'
                if dato['status_id'] == reportes_controller.anulado:
                    estado = 'anulado'

                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['cliente'], dato['tipo'], dato['subtotal'], dato['descuento'], dato['total'], dato['numero'], estado, dato['fecha']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=ventas.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_puntos = reportes_controller.lista_puntos_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_puntos': lista_puntos,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'ventas',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/ventas.html', context)


# plan pagos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def plan_pagos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        tipo = request.POST['tipo'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptPlanPagos(buffer, request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, tipo=tipo)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='plan_pagos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        # reporte de arqueo de caja

        fecha_ini = request.POST['fecha_ini'].strip()
        fecha_fin = request.POST['fecha_fin'].strip()
        ciudad = int(request.POST['ciudad'].strip())
        sucursal = int(request.POST['sucursal'].strip())
        punto = int(request.POST['punto'].strip())
        anulados = request.POST['anulados'].strip()
        tipo = request.POST['tipo'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Ventas"
            hoja.append(('Ciudad', 'Sucursal', 'Punto', 'Cliente/Almacen', 'Estado', 'Cuotas', 'Total', 'Saldo', 'Numero', 'Fecha'))
            datos_reporte = reportes_controller.datos_plan_pagos(request.user, ciudad, sucursal, punto, fecha_ini, fecha_fin, anulados, tipo=tipo)
            for dato in datos_reporte:
                estado = 'activo'
                if dato['saldo'] <= 0:
                    estado = 'concluido'
                if dato['status_id'] == reportes_controller.anulado:
                    estado = 'anulado'

                hoja.append((dato['ciudad'], dato['sucursal'], dato['punto'], dato['cliente_almacen'], estado, dato['numero_cuotas'], dato['monto_total'], dato['saldo'], dato['numero'], dato['fecha']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=plan_pagos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_ciudades = Ciudades.objects.filter(status_id=status_activo).order_by('ciudad')

    lista_puntos = reportes_controller.lista_puntos_sucursal(request.user)
    fecha_actual = get_date_show(fecha=get_date_system(), formato_ori='yyyy-mm-dd', formato='dd-MMM-yyyy')

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_puntos': lista_puntos,
        'lista_ciudades': lista_ciudades,
        'fecha_actual': fecha_actual,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'plan_pagos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/plan_pagos.html', context)


# stock productos
@user_passes_test(lambda user: get_user_permission_operation(user, settings.MOD_REPORTES, 'lista'), 'without_permission')
def stock_productos(request):
    # url main
    permisos = get_permissions_user(request.user, settings.MOD_REPORTES)

    # controler
    reportes_controller = ReportesController()
    status_activo = Status.objects.get(pk=reportes_controller.activo)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print':
        linea = request.POST['linea'].strip()
        almacen = request.POST['almacen'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            rptStockProductos(buffer, request.user, linea, almacen)

            buffer.seek(0)
            # return FileResponse(buffer, as_attachment=True, filename='hello.pdf')
            return FileResponse(buffer, filename='stock_productos.pdf')
        else:
            url = reverse('without_permission')
            return HttpResponseRedirect(url)

    if 'operation_x2' in request.POST.keys() and request.POST['operation_x2'] == 'print_excel':
        linea = request.POST['linea'].strip()
        almacen = request.POST['almacen'].strip()

        if permisos.imprimir:
            buffer = io.BytesIO()
            buffer.seek(0)

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "StockProductos"
            hoja.append(('Almacen', 'Linea', 'Producto', 'F.Elab.', 'F.Venc.', 'Lote', 'Cantidad'))
            datos_reporte = reportes_controller.datos_stock_productos(request.user, linea, almacen)

            for dato in datos_reporte:
                hoja.append((dato['almacen'], dato['linea'], dato['producto'], dato['fecha_elaboracion'], dato['fecha_vencimiento'], dato['lote'], dato['cantidad']))

            response = HttpResponse(content_type="application/msexcel")
            response["Content-Disposition"] = "attachment; filename=stock_productos.xlsx"
            wb.save(response)
            return response

    # lista de cajas de la sucursal
    lista_lineas = Lineas.objects.filter(status_id=status_activo).order_by('linea')

    # lista de almacenes
    lista_almacenes = reportes_controller.lista_almacenes_sucursal(request.user)

    # datos por defecto
    context = {
        'url_main': '',
        'js_file': reportes_controller.modulo_session,
        'lista_lineas': lista_lineas,
        'lista_almacenes': lista_almacenes,
        'permisos': permisos,

        'autenticado': 'si',

        'module_x': settings.MOD_REPORTES,
        'module_x2': '',
        'module_x3': '',

        'operation_x': 'stock_productos',
        'operation_x2': '',
        'operation_x3': '',

        'id': '',
        'id2': '',
        'id3': '',
    }
    return render(request, 'reportes/stock_productos.html', context)
